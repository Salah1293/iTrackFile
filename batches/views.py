from django.shortcuts import render, redirect
from .models import *
from django.db.models import Q, Value
from .utils import *
from .forms import *
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib import messages
from django.http import FileResponse, HttpResponse, HttpResponseRedirect
from django.urls import reverse
import os
import base64
import io
from PIL import Image
from django.http import JsonResponse
from django.template.loader import render_to_string
from users.decorators import *
from users.decorators import roles_required
from django.contrib.auth.decorators import login_required
from django.contrib.messages import error
from users.models import PvdmUsers1, Role
from django.utils.dateformat import DateFormat
from django.utils.dateformat import DateFormat
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime
from django.http import StreamingHttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

# Create your views here.


# @login_required
# @roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def landingBatches(request):
    if not request.user.is_authenticated:
        error(request, "You do not have permission to view this page.")
        return redirect('login')

    return render(request, 'batches/landing-page.html')

# general search


@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def general_search(request):

    query = request.GET.get('text', '')

    if not query:
        return render(request, 'batches/landing-page.html', {'query': query})

    all_results = generate_all_results(query)
    

    context = {'query': query, 'all_results': all_results}
    return render(request, 'batches/general-results.html', context)


# display results for individual court section
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def view_section_results(request):
    chosen_section = request.GET.get('table_name')
    query = request.GET.get('query')

    section = get_section_name(chosen_section)

    if chosen_section and query:
        all_results = generate_all_results(query, chosen_section)

        section_result = None
        for result in all_results.values():
            if result:
                section_result = result
                break

        if not section_result:
            return render(request, 'batches/general-results.html')

        resultCount = section_result.count() if query else 0

        custom_range, section_result = paginate(request, section_result)

        all_ids = []
        for obj in section_result.object_list:
            all_ids.append(obj.docid)

        template_name = f"batches/{chosen_section.lower().replace(' ', '-')
                                   }-results.html"

        context = {'section_result': section_result,
                   'custom_range': custom_range,
                   'all_ids': all_ids,
                   'resultCount': resultCount,
                   'section': section
                   }

        return render(request, template_name, context)

    else:
        return render(request, 'batches/general-results.html')



# criminal search
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def criminalSearch(request):
    return createForm(request, 'batches/criminal-search.html', CriminalForm)

# criminal display resylts


@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def criminalResults(request):

    form = CriminalForm()

    if request.method == "GET":
        form = CriminalForm(request.GET)
        if form.is_valid():
            case_number = form.cleaned_data.get('docindex1', '')
            defendant_last_name = form.cleaned_data.get('docindex6', '')
            defendant_first_name = form.cleaned_data.get('docindex7', '')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')

            query = Q()

            if case_number:
                query &= Q(docindex1=case_number)
            if start_date and end_date:
                query &= Q(docindex2__range=[start_date, end_date])
            if defendant_last_name:
                query &= Q(docindex6=defendant_last_name)
            if defendant_first_name:
                query &= Q(docindex7=defendant_first_name)

            criminal = PvdmDocs11.objects.filter(
                query) if query else PvdmDocs11.objects.none()

            resultCount = criminal.count() if query else 0

            custom_range, page_obj = paginate(request, criminal)

            for card in page_obj:
                if card.docindex2 or card.docindex20:
                    card.docindex2 = DateFormat(card.docindex2).format('Y-m-d') if card.docindex2 else None
                    card.docindex20 = DateFormat(card.docindex20).format('Y-m-d') if card.docindex20 else None

            all_ids = list(card.docid for card in page_obj if card.docid)

            context = {'section_result': page_obj,
                       'resultCount': resultCount,
                       'custom_range': custom_range,
                       'all_ids': all_ids,
                       'section': 'criminal'}
            return render(request, 'batches/criminal-results.html', context)

    context = {'section_result': PvdmDocs11.objects.none(), 'resultCount': 0}
    return redirect('criminalResults')


# civil search
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def civilSearch(request):
    return createForm(request, 'batches/civil-search.html', CivilForm)


# civil display resylts
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def civilResults(request):
    form = CivilForm()

    if request.method == 'GET':
        form = CivilForm(request.GET)
        if form.is_valid():
            case_number = form.cleaned_data.get('case_number', '')
            last_name = form.cleaned_data.get('last_name', '')
            first_name = form.cleaned_data.get('first_name', '')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')

            query = Q()

            if case_number:
                query &= Q(docindex1=case_number)
            if start_date and end_date:
                query &= Q(docindex2__range=[start_date, end_date])

            if last_name:
                query &= (Q(docindex11=last_name) | Q(docindex6=last_name) |
                          Q(docindex16=last_name))

            if first_name:
                query &= (Q(docindex12=first_name) | Q(docindex7=first_name) |
                          Q(docindex17=first_name))

            civil = PvdmDocs12.objects.filter(
                query) if query else PvdmDocs12.objects.none()

            resultCount = civil.count() if query else 0

            custom_range, page_obj = paginate(request, civil)

            for card in page_obj:
                if card.docindex2 or card.docindex20:
                    card.docindex2 = DateFormat(card.docindex2).format('Y-m-d') if card.docindex2 else None
                    card.docindex20 = DateFormat(card.docindex20).format('Y-m-d') if card.docindex20 else None

            all_ids = list(card.docid for card in page_obj if card.docid)

            context = {'section_result': page_obj,
                       'resultCount': resultCount,
                       'custom_range': custom_range,
                       'all_ids': all_ids,
                       'section': 'civil'}
            return render(request, 'batches/civil-results.html', context)

    context = {'section_result': PvdmDocs12.objects.none(), 'resultCount': 0}
    return redirect('civilResults')


# criminal cases search
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def criminalCasesSearch(request):
    return createForm(request, 'batches/criminal-cases-search.html', CriminalCasesForm)


# criminal cases display results
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def criminalCasesResults(request):
    form = CriminalCasesForm()
    if request.method == "GET":
        form = CriminalCasesForm(request.GET)
        if form.is_valid():
            date = form.cleaned_data.get('docindex1', '')
            case_number = form.cleaned_data.get('docindex2', '')
            last_Corporation = form.cleaned_data.get('docindex3', '')
            first = form.cleaned_data.get('docindex4', '')

            query = Q()

            if date:
                query &= Q(docindex1=date)
            if case_number:
                query &= Q(docindex2=case_number)
            if last_Corporation:
                query &= Q(docindex3=last_Corporation)
            if first:
                query &= Q(docindex4=first)

            criminalCases = PvdmDocs17.objects.filter(
                query) if query else PvdmDocs17.objects.none()

            resultCount = criminalCases.count() if query else 0

            custom_range, page_obj = paginate(request, criminalCases)

            for card in page_obj:
                if card.createdate :
                    card.createdate = DateFormat(card.createdate).format('Y-m-d') if card.createdate else None

            all_ids = list(card.docid for card in page_obj if card.docid)

            context = {'section_result': page_obj,
                       'resultCount': resultCount,
                       'custom_range': custom_range,
                       'all_ids': all_ids,
                       'section': 'criminal-cases'}
            return render(request, 'batches/criminal-cases-results.html', context)

    context = {'section_result': PvdmDocs17.objects.none(), 'resultCount': 0}
    return redirect('criminalCasesResults')


# criminal junevile search
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def criminalJuvenileSearch(request):
    return createForm(request, 'batches/criminal-juvenile-Search.html', CriminalJuvenileForm)


# criminal junevile display results
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def criminalJuvenileResults(request):
    form = CriminalJuvenileForm()
    if request.method == 'GET':
        form = CriminalJuvenileForm(request.GET)
        if form.is_valid():
            case_number = form.cleaned_data.get('docindex1', '')
            defendant_last_name = form.cleaned_data.get('docindex6', '')
            defendant_first_name = form.cleaned_data.get('docindex7', '')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')

            query = Q()

            if case_number:
                query &= Q(docindex1=case_number)
            if start_date and end_date:
                query &= Q(docindex2__range=[start_date, end_date])
            if defendant_last_name:
                query &= Q(docindex6=defendant_last_name)
            if defendant_first_name:
                query &= Q(docindex7=defendant_first_name)

            ciminalJuvenile = PvdmDocs13.objects.filter(
                query) if query else PvdmDocs13.objects.none()

            resultCount = ciminalJuvenile.count() if query else 0

            custom_range, page_obj = paginate(request, ciminalJuvenile)

            for card in page_obj:
                if card.docindex2 or card.docindex20:
                    card.docindex2 = DateFormat(card.docindex2).format('Y-m-d') if card.docindex2 else None
                    card.docindex20 = DateFormat(card.docindex20).format('Y-m-d') if card.docindex20 else None

            all_ids = list(card.docid for card in page_obj if card.docid)

            context = {'section_result': page_obj,
                       'resultCount': resultCount,
                       'custom_range': custom_range,
                       'all_ids': all_ids,
                       'section': 'criminal-juvenile'}
            return render(request, 'batches/criminal-juvenile-results.html', context)

    context = {'section_result': PvdmDocs13.objects.none(), 'resultCount': 0}
    return redirect('criminalJuvenileResults')


# historic index cards search
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def historicIndexCardsSearch(request):
    return createForm(request, 'batches/hisoric-index-cards-search.html', HistoricIndexCardsForm)


# historic index cards display results
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def historicIndexCardsResults(request):
    form = HistoricIndexCardsForm()
    export_url = None
    if request.method == 'GET':
        form = HistoricIndexCardsForm(request.GET)
        query_params = request.GET.dict()
        query_params.pop('csrfmiddlewaretoken', None)
        if form.is_valid():
            last_name = form.cleaned_data.get('docindex1', '')
            first_name = form.cleaned_data.get('docindex2', '')
            subject = form.cleaned_data.get('docindex3', '')
            record_source = form.cleaned_data.get('docindex4', '')
            book_record = form.cleaned_data.get('docindex5', '')
            instrument = form.cleaned_data.get('docindex9', '')
            status = form.cleaned_data.get('docindex10', '')
            owner = form.cleaned_data.get('docindex11', '')

            query = Q()

            if last_name:
                query &= Q(docindex1=last_name)
            if first_name:
                query &= Q(docindex2=first_name)
            if subject:
                query &= Q(docindex3=subject)
            if record_source:
                query &= Q(docindex4=record_source)
            if book_record:
                query &= Q(docindex5=book_record)
            if instrument:
                query &= Q(docindex9=instrument)
            if status:
                query &= Q(docindex10=status)
            if owner:
                query &= Q(docindex11=owner)

            historicIndexCards = PvdmDocs114.objects.filter(
                query) if query else PvdmDocs114.objects.none()

            resultCount = historicIndexCards.count() if query else 0

            custom_range, page_obj = paginate(
                request, historicIndexCards)
            for card in page_obj:
                if card.docindex7 :
                    card.docindex7 = DateFormat(card.docindex7).format('Y-m-d') if card.docindex7 else None

            all_ids = list(card.docid for card in page_obj if card.docid)
            query_string = '&'.join([f'{key}={value}' for key, value in query_params.items()])
            export_url = f"{reverse('export_historic_cards_table')}?{query_string}"
            is_card=True
            context = {'section_result': page_obj,
                       'resultCount': resultCount,
                       'custom_range': custom_range,
                       'all_ids': all_ids,
                       'export_url': export_url,
                       'is_card': is_card,
                       'section': 'historic-index-cards'}
            return render(request, 'batches/historic-index-cards-results.html', context)

    context = {'section_result': PvdmDocs114.objects.none(), 'export_url': export_url, 'resultCount': 0}
    return redirect('hitoricIndexCardsResults')


# historic order books search
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def historicOrderBooksSearch(request):
    return createForm(request, 'batches/historic-order-books-Search.html', HistoricOrderBooksForm)


# historic order books display results
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def historicOrderBooksResults(request):
    form = HistoricOrderBooksForm()
    export_url = None
    if request.method == 'GET':
        form = HistoricOrderBooksForm(request.GET)
        query_params = request.GET.dict()
        query_params.pop('csrfmiddlewaretoken', None)
        if form.is_valid():
            book_type_label = form.cleaned_data.get('docindex1', '')
            year = form.cleaned_data.get('docindex2', '')

            BOOK_TYPE_CHOICES = dict(HistoricOrderBooksForm.BOOK_TYPE_CHOICES)

            book_type = BOOK_TYPE_CHOICES.get(book_type_label, '')

            query = Q()

            if book_type:
                query &= Q(docindex1=book_type)
            if year:
                query &= Q(docindex2=year)

            historicOrderBooks = PvdmDocs113.objects.filter(
                query) if query else PvdmDocs113.objects.none()

            resultCount = historicOrderBooks.count() if query else 0

            custom_range, page_obj = paginate(
                request, historicOrderBooks)
            
            

            all_ids = page_obj.object_list.values_list(
                'docid', flat=True)
            query_string = '&'.join([f'{key}={value}' for key, value in query_params.items()])
            export_url = f"{reverse('export_historic_books_table')}?{query_string}"
            is_book =True
            context = {'section_result': page_obj,
                       'resultCount': resultCount,
                       'custom_range': custom_range,
                       'export_url': export_url,
                       'all_ids': all_ids,
                       'is_book' : is_book,
                       'section': 'historic-order-books'}
            return render(request, 'batches/historic-order-books-results.html', context)

    context = {'section_result': PvdmDocs113.objects.none(), 'export_url': export_url, 'resultCount': 0}
    return redirect('historicOrderBooks')


# hr search
@login_required
@roles_required('hr_manager', 'hr_staff', 'it_manager')
def hrSearch(request):
    return createForm(request, 'batches/hr-search.html', HrForm)


# hr display results
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff')
def hrResults(request):

    form = HrForm()
    if request.method == 'GET':
        form = HrForm(request.GET)
        if form.is_valid():
            last_name = form.cleaned_data.get('docindex1', '')
            first_name = form.cleaned_data.get('docindex2', '')
            ein = form.cleaned_data.get('docindex3', '')

            query = Q()

            if last_name:
                query &= Q(docindex1=last_name)
            if first_name:
                query &= Q(docindex2=first_name)
            if ein:
                query &= Q(docindex3=ein)

            try:
                login_user = PvdmUsers1.objects.get(user=request.user)
                login_user_role = login_user.role.name

                if login_user_role == 'hr_manager':
                    query &= Q(docindex6__in=['Medical', 'Personnel']) | Q(
                        docindex6=None)

                elif login_user_role == 'hr_staff':
                    query &= Q(docindex6__in=['Personnel']) | Q(docindex6=None)

            except PvdmUsers1.DoesNotExist:
                query = Q(pk=-1)

            hr = PvdmDocs15.objects.filter(
                query) if query else PvdmDocs15.objects.none()

            resultCount = hr.count() if query else 0

            custom_range, page_obj = paginate(request, hr)

            for card in page_obj:
                if card.createdate:
                    card.createdate = DateFormat(card.createdate).format('Y-m-d') if card.createdate else None

            all_ids = list(card.docid for card in page_obj if card.docid)

            context = {'section_result': page_obj,
                       'resultCount': resultCount,
                       'custom_range': custom_range,
                       'all_ids': all_ids,
                       'section': 'hr'}
            return render(request, 'batches/hr-results.html', context)

    context = {'section_result': PvdmDocs15.objects.none(), 'resultCount': 0}
    return render(request, 'batches/hr-results.html', context)


# bond books search
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def bondBooksSearch(request):
    return createForm(request, 'batches/bond-books-search.html', BondBooksForm)

# bond books display results


@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def bondBooksResults(request):

    form = BondBooksForm()
    bondBooks = PvdmDocs116.objects.none()
    all_ids = []
    resultCount = 0

    if request.method == 'GET':
        form = BondBooksForm(request.GET)
        if form.is_valid():
            book = form.cleaned_data.get('docindex1', '')
            page = form.cleaned_data.get('docindex2', '')

            query = Q()

            if book:
                query &= Q(docindex1=book)
            if page:
                query &= Q(docindex2=page)

            bondBooks = PvdmDocs116.objects.filter(
                query) if query else PvdmDocs116.objects.none()

            resultCount = bondBooks.count() if query else 0

            custom_range, page_obj = paginate(request, bondBooks)

            for card in page_obj:
                if card.createdate :
                    card.createdate = DateFormat(card.createdate).format('Y-m-d') if card.createdate else None

            all_ids = list(card.docid for card in page_obj if card.docid)


            context = {'section_result': page_obj,
                       'resultCount': resultCount,
                       'custom_range': custom_range,
                       'all_ids': all_ids,
                       'section': 'bond'}
            return render(request, 'batches/bond-books-results.html', context)

    context = {'section_result': PvdmDocs116.objects.none(),
               'resultCount': resultCount}
    return render(request, 'batches/bond-books-results.html', context)


# charters search
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def chartersSearch(request):
    return createForm(request, 'batches/charters-search.html', ChartersForm)


# charters display results
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def chartersResults(request):
    form = ChartersForm()
    if request.method == 'GET':
        form = ChartersForm(request.GET)
        if form.is_valid():
            charter_name = form.cleaned_data.get('docindex1', '')
            book = form.cleaned_data.get('docindex2', '')
            page = form.cleaned_data.get('docindex3', '')

            query = Q()

            if charter_name:
                query &= Q(docindex1=charter_name)
            if book:
                query &= Q(docindex2=book)
            if page:
                query &= Q(docindex3=page)

            charters = PvdmDocs19.objects.filter(
                query) if query else PvdmDocs19.objects.none()

            resultCount = charters.count() if query else 0

            custom_range, page_obj = paginate(request, charters)

            for card in page_obj:
                if card.createdate :
                    card.createdate = DateFormat(card.createdate).format('Y-m-d') if card.createdate else None

            all_ids = list(card.docid for card in page_obj if card.docid)

            context = {'section_result': page_obj,
                       'resultCount': resultCount,
                       'custom_range': custom_range,
                       'all_ids': all_ids,
                       'section': 'charters'}
            return render(request, 'batches/charters-results.html', context)

    context = {'section_result': PvdmDocs19.objects.none(), 'resultCount': 0}
    return render(request, 'batches/charters-results.html', context)


# concealed weapons search
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def ConcealedWeaponsSearch(request):
    return createForm(request, 'batches/concealed-weapons-search.html', ConcealedWeaponsForm)


# concealed weapons display results
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def ConcealedWeaponsResults(request):

    form = ConcealedWeaponsForm()
    if request.method == 'GET':
        form = ConcealedWeaponsForm(request.GET)
        if form.is_valid():
            case_number = form.cleaned_data.get('docindex1', '')
            subject_company = form.cleaned_data.get('docindex7', '')
            subject_last_name = form.cleaned_data.get('docindex8', '')
            subject_first_name = form.cleaned_data.get('docindex9', '')

            query = Q()

            if case_number:
                query &= Q(docindex1=case_number)
            if subject_company:
                query &= Q(docindex7=subject_company)
            if subject_last_name:
                query &= Q(docindex8=subject_last_name)
            if subject_first_name:
                query &= Q(docindex9=subject_first_name)

            concealedWeapons = PvdmDocs112.objects.filter(
                query) if query else PvdmDocs112.objects.none()

            resultCount = concealedWeapons.count() if query else 0

            custom_range, page_obj = paginate(
                request, concealedWeapons)
            
            for card in page_obj:
                if card.createdate or card.docindex2 or card.docindex6:
                    card.createdate = DateFormat(card.createdate).format('Y-m-d') if card.createdate else None
                    card.docindex2 = DateFormat(card.docindex2).format('Y-m-d') if card.docindex2 else None
                    card.docindex6 = DateFormat(card.docindex6).format('Y-m-d') if card.docindex6 else None

            all_ids = list(card.docid for card in page_obj if card.docid)


            context = {'section_result': page_obj,
                       'resultCount': resultCount,
                       'custom_range': custom_range,
                       'all_ids': all_ids,
                       'section': 'concealed-weapons'}
            return render(request, 'batches/concealed-weapons-results.html', context)

    context = {'section_result': PvdmDocs112.objects.none(), 'resultCount': 0}
    return render(request, 'batches/concealed-weapons-results.html', context)


# indictmenys search
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def indictmentsSearch(request):
    return createForm(request, 'batches/indictments-search.html', IndictmentsForm)


# indictmenys display results
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def indictmentsResults(request):
    form = IndictmentsForm()
    if request.method == 'GET':
        form = IndictmentsForm(request.GET)
        if form.is_valid():
            case_number = form.cleaned_data.get('docindex1', '')
            defendant_first_name = form.cleaned_data.get('docindex3', '')
            defendant_last_name = form.cleaned_data.get('docindex5', '')

            query = Q()

            if case_number:
                query &= Q(docindex1=case_number)
            if defendant_first_name:
                query &= Q(docindex3=defendant_first_name)
            if defendant_last_name:
                query &= Q(docindex5=defendant_last_name)

            indictments = PvdmDocs110.objects.filter(
                query) if query else PvdmDocs110.objects.none()

            resultCount = indictments.count() if query else 0

            custom_range, page_obj = paginate(request, indictments)

            for card in page_obj:
                if card.createdate or card.docindex8 or card.docindex9:
                    card.createdate = DateFormat(card.createdate).format('Y-m-d') if card.createdate else None
                    card.docindex8 = DateFormat(card.docindex8).format('Y-m-d') if card.docindex8 else None
                    card.docindex9 = DateFormat(card.docindex9).format('Y-m-d') if card.docindex9 else None

            all_ids = list(card.docid for card in page_obj if card.docid)


            context = {'section_result': page_obj,
                       'resultCount': resultCount,
                       'custom_range': custom_range,
                       'all_ids': all_ids,
                       'section': 'indictments'}
            return render(request, 'batches/indictments-results.html', context)

    context = {'section_result': PvdmDocs110.objects.none(), 'resultCount': 0}
    return render(request, 'batches/indictments-results.html', context)


# law&chancery search
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def lawChancerySearch(request):
    return createForm(request, 'batches/law-chancery-search.html', LawChanceryForm)


# law&chancery display results
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def lawChanceryResults(request):
    form = LawChanceryForm()
    if request.method == 'GET':
        form = LawChanceryForm(request.GET)
        if form.is_valid():
            date = form.cleaned_data.get('docindex1', '')
            law_Chancery = form.cleaned_data.get('docindex2', '')

            query = Q()

            if date:
                query &= Q(docindex1=date)
            if law_Chancery:
                query &= Q(docindex2=law_Chancery)

            lawChancery = PvdmDocs16.objects.filter(
                query) if query else PvdmDocs16.objects.none()

            resultCount = lawChancery.count() if query else 0

            custom_range, page_obj = paginate(request, lawChancery)

            for card in page_obj:
                if card.createdate :
                    card.createdate = DateFormat(card.createdate).format('Y-m-d') if card.createdate else None

            all_ids = list(card.docid for card in page_obj if card.docid)


            context = {'section_result': page_obj,
                       'resultCount': resultCount,
                       'custom_range': custom_range,
                       'all_ids': all_ids,
                       'section': 'law-chancery'}
            return render(request, 'batches/law-chancery-results.html', context)

    context = {'section_result': PvdmDocs16.objects.none(), 'resultCount': 0}
    return render(request, 'batches/law-chancery-results.html', context)


# destruction orders search
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def destructionOrdersSearch(request):
    return createForm(request, 'batches/destruction-orders-search.html', DestructionOrdersForm)


# destruction orders display results
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def destructionOrdersResults(request):
    form = DestructionOrdersForm()
    if request.method == 'GET':
        form = DestructionOrdersForm(request.GET)
        if form.is_valid():
            order_type = form.cleaned_data.get('docindex1', '')
            order_date = form.cleaned_data.get('docindex2', '')

            query = Q()

            if order_type:
                query &= Q(docindex1=order_type)
            if order_date:
                query &= Q(docindex2=order_date)

            destructionOrders = PvdmDocs115.objects.filter(
                query) if query else PvdmDocs115.objects.none()

            resultCount = destructionOrders.count() if query else 0

            custom_range, page_obj = paginate(
                request, destructionOrders)
            
            for card in page_obj:
                if  card.createdate or card.docindex2 or card.docindex or card.docindex4 or card.docindex5:
                    card.createdate = DateFormat(card.createdate).format('Y-m-d') if card.createdate else None
                    card.docindex2 = DateFormat(card.docindex2).format('Y-m-d') if card.docindex2 else None
                    card.docindex3 = DateFormat(card.docindex3).format('Y-m-d') if card.docindex3 else None
                    card.docindex4 = DateFormat(card.docindex4).format('Y-m-d') if card.docindex4 else None
                    card.docindex5 = DateFormat(card.docindex5).format('Y-m-d') if card.docindex5 else None

            all_ids = list(card.docid for card in page_obj if card.docid)

            context = {'section_result': page_obj,
                       'resultCount': resultCount,
                       'custom_range': custom_range,
                       'all_ids': all_ids,
                       'section': 'destruction-orders'}
            return render(request, 'batches/destruction-orders-results.html', context)

    context = {'section_result': PvdmDocs115.objects.none(), 'resultCount': 0}
    return render(request, 'batches/destruction-orders-results.html', context)


# adoption search
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def adoptionSearch(request):
    return createForm(request, 'batches/adoption-search.html', AdoptionForm)


# adoption display results
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def adoptionResults(request):
    form = AdoptionForm()
    if request.method == 'GET':
        form = AdoptionForm(request.GET)
        if form.is_valid():
            case_number = form.cleaned_data.get('docindex1', '')
            plaintiff_last_name = form.cleaned_data.get('docindex6', '')
            plaintiff_first_name = form.cleaned_data.get('docindex7', '')
            subject_last_name = form.cleaned_data.get('docindex16', '')
            subject_first_name = form.cleaned_data.get('docindex17', '')

            query = Q()

            if case_number:
                query &= Q(docindex1=case_number)
            if plaintiff_last_name:
                query &= Q(docindex6=plaintiff_last_name)
            if plaintiff_first_name:
                query &= Q(docindex7=plaintiff_first_name)
            if subject_last_name:
                query &= Q(docindex16=subject_last_name)
            if subject_first_name:
                query &= Q(docindex17=subject_first_name)

            adoption = PvdmDocs14.objects.filter(
                query) if query else PvdmDocs14.objects.none()

            resultCount = adoption.count() if query else 0

            custom_range, page_obj = paginate(request, adoption)

            for card in page_obj:
                if  card.createdate or card.docindex2 or card.docindex20:
                    card.createdate = DateFormat(card.createdate).format('Y-m-d') if card.createdate else None
                    card.docindex2 = DateFormat(card.docindex2).format('Y-m-d') if card.docindex2 else None
                    card.docindex20 = DateFormat(card.docindex20).format('Y-m-d') if card.docindex20 else None

            all_ids = list(card.docid for card in page_obj if card.docid)

            context = {'section_result': page_obj,
                       'resultCount': resultCount,
                       'custom_range': custom_range,
                       'all_ids': all_ids,
                       'section': 'adoption'}
            return render(request, 'batches/adoption-results.html', context)

    context = {'section_result': PvdmDocs14.objects.none(), 'resultCount': 0}
    return render(request, 'batches/adoption-results.html', context)


# clerk orders search
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def clerkOrdersSearch(request):
    return createForm(request, 'batches/clerk-orders-search.html', ClerkOrdersForm)


# clerk orders display results
@login_required
@roles_required('admin', 'hr_manager', 'hr_staff', 'general_staff', 'historic_staff')
def clerkOrdersResults(request):
    form = ClerkOrdersForm()
    if request.method == 'GET':
        form = ClerkOrdersForm(request.GET)
        if form.is_valid():
            case_number = form.cleaned_data.get('docindex1', '')
            plaintiff_company = form.cleaned_data.get('docindex8', '')
            plaintiff_last_name = form.cleaned_data.get('docindex9', '')
            plaintiff_first_name = form.cleaned_data.get('docindex10', '')
            defendant_company = form.cleaned_data.get('docindex13', '')
            defendant_last_name = form.cleaned_data.get('docindex14', '')
            defendant_first_name = form.cleaned_data.get('docindex15', '')
            subject_company = form.cleaned_data.get('docindex18', '')
            subject_last_name = form.cleaned_data.get('docindex19', '')
            subject_first_name = form.cleaned_data.get('docindex20', '')

            query = Q()

            if case_number:
                query &= Q(docindex1=case_number)
            if plaintiff_company:
                query &= Q(docindex8=plaintiff_company)
            if plaintiff_last_name:
                query &= Q(docindex9=plaintiff_last_name)
            if plaintiff_first_name:
                query &= Q(docindex10=plaintiff_first_name)
            if defendant_company:
                query &= Q(docindex13=defendant_company)
            if defendant_last_name:
                query &= Q(docindex14=defendant_last_name)
            if defendant_first_name:
                query &= Q(docindex15=defendant_first_name)
            if subject_company:
                query &= Q(docindex18=subject_company)
            if subject_last_name:
                query &= Q(docindex19=subject_last_name)
            if subject_first_name:
                query &= Q(docindex20=subject_first_name)

            clerkOrder = PvdmDocs18.objects.filter(
                query) if query else PvdmDocs18.objects.none()

            resultCount = clerkOrder.count() if query else 0

            custom_range, page_obj = paginate(request, clerkOrder)

            for card in page_obj:
                if  card.createdate or card.docindex2 or card.docindex23:
                    card.createdate = DateFormat(card.createdate).format('Y-m-d') if card.createdate else None
                    card.docindex2 = DateFormat(card.docindex2).format('Y-m-d') if card.docindex2 else None
                    card.docindex23 = DateFormat(card.docindex23).format('Y-m-d') if card.docindex23 else None

            all_ids = list(card.docid for card in page_obj if card.docid)

            context = {'section_result': page_obj,
                       'resultCount': resultCount,
                       'custom_range': custom_range,
                       'all_ids': all_ids,
                       'section': 'clerk-orders'}
            return render(request, 'batches/clerk-orders-results.html', context)

    context = {'section_result': PvdmDocs18.objects.none(), 'resultCount': 0}
    return render(request, 'batches/clerk-orders-results.html', context)


def single_image_view(request, section, pk):

    model_mapping = {
        'bond': (PvdmDocs116, PvdmObjs116, UpdateBondBooks),
        'criminal': (PvdmDocs11, PvdmObjs11, UpdateCriminal),
        'civil': (PvdmDocs12, PvdmObjs12, UpdateCivil),
        'criminal-cases': (PvdmDocs17, PvdmObjs17, UpdateCriminalCases),
        'criminal-juvenile': (PvdmDocs13, PvdmObjs13, UpdateCriminalJuvenile),
        'historic-index-cards': (PvdmDocs114, PvdmObjs114, UpdateHistoricIndexCards),
        'historic-order-books': (PvdmDocs113, PvdmObjs113, UpdateHistoricOrderBooks),
        'hr': (PvdmDocs15, PvdmObjs15, UpdateHr),
        'charters': (PvdmDocs19, PvdmObjs19, UpdateCharters),
        'concealed-weapons': (PvdmDocs112, PvdmObjs112, UpdateConcealedWeapons),
        'indictments': (PvdmDocs110, PvdmObjs110, UpdateIndictments),
        'law-chancery': (PvdmDocs16, PvdmObjs16, UpdateLawChancery),
        'destruction-orders': (PvdmDocs115, PvdmObjs115, UpdateDestructionOrders),
        'adoption': (PvdmDocs14, PvdmObjs14, UpdateAdoption),
        'clerk-orders': (PvdmDocs18, PvdmObjs18, UpdateClerkOrders),
    }

    if section not in model_mapping:
        raise ValueError("Invalid section")

    doc_model, obj_model, card_form = model_mapping[section]

    return singleImageView(request, pk, doc_model, obj_model, card_form, section)





#generates excel file for historic index cards
def export_historic_cards_table(request):
    query_params = request.GET.dict()
    query_params.pop('csrfmiddlewaretoken', None)
    page_number = query_params.pop('page', None)

    query = Q()
    for key, value in query_params.items():
        if value:
            query &= Q(**{key: value})

    historicIndexCards = PvdmDocs114.objects.filter(query).only(
        'docindex1', 'docindex2', 'docindex3', 'docindex10',
        'docindex11', 'docindex9', 'docindex4', 'docindex5',
        'docindex6', 'docindex7', 'docindex12', 'docindex8'
    ).iterator(chunk_size=1000)

    headers = [
        'Last Name', 'First Name', 'Subject', 'Status', 'Owner',
        'Instrument', 'Record Source', 'Book Record', 'Page', 'Date',
        'Dates Before 1753', 'Comments'
    ]

    def generate_excel():
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Historic Index Cards'

        worksheet.append(headers)

        for element in historicIndexCards:
            row = [
                element.docindex1, element.docindex2, element.docindex3, element.docindex10,
                element.docindex11, element.docindex9, element.docindex4, element.docindex5,
                element.docindex6, element.docindex7, element.docindex12, element.docindex8
            ]
            row = [field.strftime('%Y-%m-%d') if isinstance(field, datetime) else field for field in row]
            worksheet.append(row)

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        yield save_virtual_workbook(workbook)

    current_datetime = datetime.now().strftime('(%Y-%m-%d)-(%H:%M %p)')
    response = StreamingHttpResponse(
        generate_excel(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="historic-index-cards-{current_datetime}.xlsx"'
    return response



#Generates excel file for historic order books
def export_historic_books_table(request):
    BOOK_TYPE_CHOICES = dict([
        ('bond_book_1', 'Bond Book 1'),
        ('court_order_superior', 'Court Order Superior'),
        ('court_quarterly_sessions', 'Court Quarterly Sessions'),
        ('estray_book', 'Estray Book'),
        ('land_causes_1', 'Land Causes 1'),
        ('land_causes_2', 'Land Causes 2'),
        ('land_records_long_standing', 'Land Records Long Standing'),
        ('minute_book', 'Minute Book'),
        ('ordinary_bond_book', 'Ordinary Bond Book'),
        ('quite_rents', 'Quite Rents'),
        ('reg_free_negroes_val_2', 'Reg Free Negroes Val 2'),
        ('reg_free_negroes_val_3', 'Reg Free Negroes Val 3'),
        ('roads', 'Roads'),
        ('surveys', 'Surveys'),
    ])

    query_params = request.GET.dict()
    docindex1_key = query_params.get('docindex1', '').strip()
    docindex1_value = BOOK_TYPE_CHOICES.get(docindex1_key, None)
    year = query_params.get('docindex2', '').strip()

    query = Q()
    if docindex1_value:
        query &= Q(docindex1=docindex1_value)
    if year:
        query &= Q(docindex2=year)

    historic_books = PvdmDocs113.objects.filter(query).only(
        'docindex1', 'docindex2', 'docindex3', 'docindex4').iterator(chunk_size=1000)
    
    headers = ['Book Type', 'Year', 'Page A', 'Page B']

    def generate_excel():
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Historic Order Books'

        
        worksheet.append(headers)

        for book in historic_books:
            row = [book.docindex1, book.docindex2, book.docindex3, book.docindex4]
            row = [field.strftime('%Y-%m-%d') if isinstance(field, datetime) else field for field in row]
            worksheet.append(row)

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        yield save_virtual_workbook(workbook)

    current_datetime = datetime.now().strftime('(%Y-%m-%d)-(%H:%M %p)')
    response = StreamingHttpResponse(generate_excel(),
                                     content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="historic-order-books-{current_datetime}.xlsx"'
    return response


